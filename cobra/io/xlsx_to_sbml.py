


from openpyxl import load_workbook
from libsbml import *



def get_libsbml_document_from_xls(model_xlsx,model_id,sbml_level=2,sbml_version=4):
	###################################
	rxn_headers='Rxn name|Rxn description|Formula|Gene-reaction association|\
Genes|Proteins|Subsystem|Reversible|LB|UB|Objective|\
Confidence Score|EC Number|Notes|References'.split('|')
	met_headers='Metabolite name|Metabolite description|Metabolite neutral formula|\
Metabolite charged formula|Metabolite charge|Metabolite Compartment|Metabolite KEGGID|\
Metabolite PubChemID|Metabolite CheBI ID|Metabolite Inchi String|Metabolite Smile'.split('|')
	###################################
	# Required to allow BiGG to comply with SBML, used for adding species, reactants/products
	#
	def comply_met_with_SBML(met_id):
		met_id='M_'+met_id.replace('[','_').replace(']','')\
				.replace('-','_').replace(',','_').replace("'",'_')
		return met_id
	###################################
	# Required to parse formula string into stoich and met_id
	#
	def parse_stoich_met(stoich_metabolite):
		# remove white space at the end of metabolite
		while stoich_metabolite[-1] in ' ':
			stoich_metabolite=stoich_metabolite[:-1]
		if ' ' in stoich_metabolite:
			stoich,met_id=stoich_metabolite.split(' ')
		else:
			stoich='1.0'
			met_id=stoich_metabolite
		# necessary to remove features not compatible with SBML
		met_id=comply_met_with_SBML(met_id)
		return stoich,met_id
	wb = load_workbook(filename = model_xlsx )
	######################################################
	# Check to ensure headers in xlsx match expectd format
	#
	if 'reactions' not in wb.sheetnames and 'metabolites' not in wb.sheetnames:
		print('missing reaction and/or metabolites sheet')
		quit()
	######################################################
	ws=wb['reactions']
	mismatching_rxn_columns=[cell.value \
			for header,col in zip(rxn_headers,ws.iter_cols(min_row=1, max_col=15, max_row=1)) \
			for cell in col if header != cell.value]
	if len(mismatching_rxn_columns)>0:
		print('mismatching reaction headers')
		print(mismatching_rxn_columns)
		quit()
	######################################################
	ws=wb['metabolites']
	mismatching_mets_columns=[cell.value \
			for header,col in zip(met_headers,ws.iter_cols(min_row=1, max_col=11, max_row=1)) \
			for cell in col if header != cell.value]
	if len(mismatching_mets_columns)>0:
		print('mismatching metabolite headers')
		print(mismatching_mets_columns)
		quit()
	print('spreadsheet contains correct sheets/headers')
	###################################
	# Specify SBML version
	#
	document = SBMLDocument(sbml_level, sbml_version)
	model = document.createModel()
	###################################
	# Set model name
	#
	model.setId(model_id)
	#
	###################################
	# Create unit definitions
	#
	flux = model.createUnitDefinition()
	flux.setId('mmol_per_gDW_per_hr')
	#
	###################################
	unit = flux.createUnit()
	unit.setKind(UNIT_KIND_MOLE)
	unit.setExponent(1)
	unit.setScale(-3)
	###################################
	unit = flux.createUnit()
	unit.setKind(UNIT_KIND_GRAM)
	unit.setExponent(-1)
	unit.setScale(1)
	###################################
	unit = flux.createUnit()
	unit.setKind(UNIT_KIND_SECOND)
	unit.setExponent(-1)
	unit.setScale(1)
	unit.setMultiplier(1.0 / 60 / 60)
	###################################
	# Create compartments, name is set to id by default
	#
	compartments=list(   set(  [   cell.value.split('[')[1].split(']')[0].encode()  \
		for col in wb['metabolites'].iter_rows(min_row=2, max_col=14) for cell in col[:1]]  )  )
	for compartment in compartments:
		c1 = model.createCompartment()
		c1.setId(compartment)
		c1.setName(compartment)
		#c1.setConstant(True)
		#c1.setSize(1)
		#c1.setSpatialDimensions(3)
		#c1.setUnits('litre')
	###################################
	# Create species
	# Currently only id and name are stored in xml file
	# All other information stored in notes
	for row in wb['metabolites'].iter_rows(min_row=2, max_col=11):
		# load row into dictionary
		met_information ={met_header:cell.value.encode() if isinstance(cell.value,unicode) \
							else cell.value\
							for met_header,cell in zip(met_headers,row)}
		print(met_information['Metabolite name'])
		s1 = model.createSpecies()
		# remove BOP compartment for SBML format
		s1.setId(comply_met_with_SBML(met_information['Metabolite name']))
		s1.setName(met_information['Metabolite description'])
		if met_information['Metabolite Compartment'] is None:
			met_compartment=met_information['Metabolite name'].split('[')[1].split(']')[0]
			s1.setCompartment(met_compartment)
		else:
			s1.setCompartment(met_information['Metabolite Compartment'])
		s1.setConstant(True)
		#s1.setInitialAmount(0)	
		notes_list=['<p>'+met_header+': '+str(met_information[met_header])+'</p>'\
						for met_header in met_headers[2:] if met_information[met_header] is not None]
		s1.appendNotes('<notes><body xmlns="http://www.w3.org/1999/xhtml">'+\
					''.join(notes_list)+\
					'</body></notes>')
	###################################
	# Parse reactions
	# 
	for row in wb['reactions'].iter_rows(min_row=2, max_col=15):
		rxn_information={rxn_header:cell.value.encode() if isinstance(cell.value,unicode) \
							 else cell.value  for rxn_header,cell in zip(rxn_headers,row)}
		print(rxn_information['Rxn name'])
		# consolidate into one variable
		# add reaction
		print 'create reaction'
		r1 = model.createReaction()
		r1.setId('R_'+rxn_information['Rxn name'].replace('-','_').replace('(','_').replace(')',''))
		r1.setName(rxn_information['Rxn description'])
		# COBRA reversibility does not match SBML reversibility definition?
		if rxn_information['Reversible']==1:
			r1.setReversible(True)
		elif rxn_information['Reversible']==0:
			r1.setReversible(False)
		r1.setFast(False)
		print 'add notes'
		###################################
		# Add notes
		#
		notes_list=['<p>'+rxn_header+': '+str(rxn_information[rxn_header])+'</p>'\
						 for rxn_header in rxn_headers[3:7]+rxn_headers[11:]\
						  if rxn_information[rxn_header] is not None]
		r1.appendNotes('<notes><body xmlns="http://www.w3.org/1999/xhtml">'+\
					''.join(notes_list)+\
					'</body></notes>')
		###################################
		# Parse formula into reactants and products
		#
		rxn_formula=rxn_information['Formula']
		if '->' in rxn_formula:
			if len(filter(None,rxn_formula.split('->')))==1:
				delim=' ->'
			else:
				delim=' -> '
		elif '<=>' in rxn_formula:
			if len(filter(None,rxn_formula.split('<=>')))==1:
				delim=' <=>'
			else:
				delim=' <=> '
		reactants=rxn_formula.split(delim)[0].split(' + ')
		if len(rxn_formula.split(delim))>1:
			products=rxn_formula.split(delim)[1].split(' + ')
		else:
			products=[]
		###################################
		print('add reactants')
		for reactant in reactants:
			print reactant
			stoich,met_id=parse_stoich_met(reactant)
			species_ref1 = r1.createReactant()
			species_ref1.setSpecies(met_id)
			species_ref1.setStoichiometry(float(stoich))
			species_ref1.setConstant(True)
		print('add products')
		# filter out empty products
		# A -> 
		for product in list(filter(None,products)):
			print product
			stoich,met_id=parse_stoich_met(product)
			species_ref2 = r1.createProduct()
			species_ref2.setSpecies(met_id)
			species_ref2.setStoichiometry(float(stoich))
			species_ref2.setConstant(True)
		###################################
		# Add kinietic information
		#
		kinetic_law = r1.createKineticLaw()
		k = kinetic_law.createParameter()
		k.setId('LOWER_BOUND')
		k.setValue(rxn_information['LB'])
		k.setUnits('mmol_per_gDW_per_hr')
		k = kinetic_law.createParameter()
		k.setId('UPPER_BOUND')
		k.setValue(rxn_information['UB'])
		k.setUnits('mmol_per_gDW_per_hr')
		k = kinetic_law.createParameter()
		k.setId('OBJECTIVE_COEFFICIENT')
		k.setValue(rxn_information['Objective'])
		k.setUnits('mmol_per_gDW_per_hr')
		k = kinetic_law.createParameter()
		k.setId('FLUX_VALUE')
		k.setValue(0)
		k.setUnits('mmol_per_gDW_per_hr')
	writeSBML(document,model_id+'.xml')






def get_mets_lists(model_xlsx):
	###################################
	# Returns list of metabolites in reactions and metabolites
	#
	rxn_headers='Rxn name|Rxn description|Formula|Gene-reaction association|\
Genes|Proteins|Subsystem|Reversible|LB|UB|Objective|\
Confidence Score|EC Number|Notes|References'.split('|')
	met_headers='Metabolite name|Metabolite description|Metabolite neutral formula|\
Metabolite charged formula|Metabolite charge|Metabolite Compartment|Metabolite KEGGID|\
Metabolite PubChemID|Metabolite CheBI ID|Metabolite Inchi String|Metabolite Smile'.split('|')
	###################################
	# Required to allow BiGG to comply with SBML, used for adding species, reactants/products
	#
	def comply_met_with_SBML(met_id):
		met_id='M_'+met_id.replace('[','_').replace(']','')\
				.replace('-','_').replace(',','_').replace("'",'_')
		return met_id
	###################################
	# Required to parse formula string into stoich and met_id
	#
	def parse_stoich_met(stoich_metabolite):
		# remove white space at the end of metabolite
		while stoich_metabolite[-1] in ' ':
			stoich_metabolite=stoich_metabolite[:-1]
		if ' ' in stoich_metabolite:
			stoich,met_id=stoich_metabolite.split(' ')
		else:
			stoich='1.0'
			met_id=stoich_metabolite
		# necessary to remove features not compatible with SBML
		met_id=comply_met_with_SBML(met_id)
		return stoich,met_id
	wb = load_workbook(filename = model_xlsx )
	######################################################
	# Check to ensure headers in xlsx match expectd format
	#
	if 'reactions' not in wb.sheetnames and 'metabolites' not in wb.sheetnames:
		print('missing reaction and/or metabolites sheet')
		quit()
	######################################################
	ws=wb['reactions']
	mismatching_rxn_columns=[cell.value \
			for header,col in zip(rxn_headers,ws.iter_cols(min_row=1, max_col=15, max_row=1)) \
			for cell in col if header != cell.value]
	if len(mismatching_rxn_columns)>0:
		print('mismatching reaction headers')
		print(mismatching_rxn_columns)
		quit()
	######################################################
	ws=wb['metabolites']
	mismatching_mets_columns=[cell.value \
			for header,col in zip(met_headers,ws.iter_cols(min_row=1, max_col=11, max_row=1)) \
			for cell in col if header != cell.value]
	if len(mismatching_mets_columns)>0:
		print('mismatching metabolite headers')
		print(mismatching_mets_columns)
		quit()
	print('spreadsheet contains correct sheets/headers')
	###################################
	# Create species
	# Currently only id and name are stored in xml file
	# All other information stored in notes
	output_metabolites=[]
	for row in wb['metabolites'].iter_rows(min_row=2, max_col=11):
		# load row into dictionary
		met_information ={met_header:cell.value.encode() if isinstance(cell.value,unicode) \
							else cell.value\
							for met_header,cell in zip(met_headers,row)}
		print(met_information['Metabolite name'])
		# remove BOP compartment for SBML format
		output_metabolites.append(comply_met_with_SBML(met_information['Metabolite name']))
	###################################
	# Parse reactions
	#
	output_reaction_metabolites=[]
	for row in wb['reactions'].iter_rows(min_row=2, max_col=15):
		rxn_information={rxn_header:cell.value.encode() if isinstance(cell.value,unicode) \
							 else cell.value  for rxn_header,cell in zip(rxn_headers,row)}
		rxn_formula=rxn_information['Formula']
		if '->' in rxn_formula:
			if len(filter(None,rxn_formula.split('->')))==1:
				delim=' ->'
			else:
				delim=' -> '
		elif '<=>' in rxn_formula:
			if len(filter(None,rxn_formula.split('<=>')))==1:
				delim=' <=>'
			else:
				delim=' <=> '
		reactants=rxn_formula.split(delim)[0].split(' + ')
		if len(rxn_formula.split(delim))>1:
			products=rxn_formula.split(delim)[1].split(' + ')
		else:
			products=[]
		###################################
		print('add reactants')
		for reactant in reactants:
			print reactant
			stoich,met_id=parse_stoich_met(reactant)
			output_reaction_metabolites.append(met_id)
		print('add products')
		# filter out empty products
		# A -> 
		for product in list(filter(None,products)):
			print product
			stoich,met_id=parse_stoich_met(product)
			output_reaction_metabolites.append(met_id)
	return output_metabolites,output_reaction_metabolites






model_id='yli'
model_xlsx='yli.xlsx'
sbml_level=2
sbml_version=4


mets_in_metabolites,mets_in_reactions=get_mets_lists(model_xlsx)

missing_mets_from_reactions=set(mets_in_reactions)-set(mets_in_metabolites)

missing_mets_in_reactions=set(mets_in_metabolites)-set(mets_in_reactions)

if len(missing_mets_from_reactions)==0:
	get_libsbml_document_from_xls(model_xlsx,model_id,sbml_level=2,sbml_version=4)
else:
	print('metabolites in reactions tab missing in metabolites tab')


